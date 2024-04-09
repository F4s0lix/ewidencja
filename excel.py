import pandas as pd
import os
import shutil
from openpyxl.styles import PatternFill, Border, Side
from tqdm import tqdm

class ExcelHandler:
    def __init__(self, excel_filepath: str, output_path: str = 'output.xlsx') -> None:
        if not os.path.exists(excel_filepath):
            raise FileNotFoundError(f'Nie mozna znalezc pliku: {excel_filepath}')
        self.excel_filepath = excel_filepath
        self.csv_dir = 'csv_data/'
        self.output_path = output_path

        self.CELL_COLOR = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.BORDER = Side(style='thin', color='000000')
        self.BORDER_STYLE = Border(left=self.BORDER, right=self.BORDER, top=self.BORDER, bottom=self.BORDER)
        
    def _get_data_to_csv_files(self, column_pjo: str = 'pjo', column_desc: str= 'opis', column_income: str = 'przychód', column_expenditure: str = 'rozchód') -> None:
        """
            funkcja z pliku podanego przy inicjacji klasy tworzy zestaw plikow w folderze csv_dir, na ktorych program potem pracuje
            column_X - nazwa kolumny w Excelu, ktora zawiera dana rzecz UWAGA: plik musi miec tylko jeden arkusz
        """
        print('Zaczynam Tworzyc pliki pomocnicze...')
        print(f'\tSprawdzam czy istnieje {self.csv_dir}...')
        if os.path.exists(self.csv_dir):
            print(f'\tUsuwam {self.csv_dir}')
            shutil.rmtree(self.csv_dir)
        print(f'\tTworze nowy foler {self.csv_dir}...')
        os.mkdir(self.csv_dir)
        
        print(f'\tOdczytuje dane z pliku {self.excel_filepath}...')
        DATA = pd.read_excel(self.excel_filepath)
        print('\tTworze liste arkuszy...')
        PJO = list(filter(lambda x: isinstance(x, str), set(DATA[column_pjo])))
        PJO.remove('nie kopiować')
        PJO_DICT = {j: {} for j in PJO}

        print('\tTworze zestaw rozchodu i dochodu...')
        for _, row in tqdm(DATA.iterrows(), desc='\tPrzetwarzam...', ncols=len(PJO), unit=' kolumn'):
            income = row[column_income]
            expenditure = row[column_expenditure]
            
            if row[column_pjo] in PJO:
                PJO_DICT[row[column_pjo]][row[column_desc]] = (income, expenditure)
        
        print(f'\tTworze osobne pliki pomocnicze dla: {", ".join(PJO)}')
        for p in tqdm(PJO, desc='\tPliki pomocnicze: ', unit=' jednostek'):
            try:
                with open(f'{self.csv_dir}/{p}.csv', 'a') as file:
                    file.write('opis;przychod;rozchod\n') 
                    for desc, money in PJO_DICT[p].items():
                        income, expend = money
                        file.write(f'{desc};{income:.2f};{expend:.2f}\n')
            except Exception as e:
                print(f'Nie mozna stworzyc plikow pomocniczych.\nBlad: {e}')
        print('Udalo sie utworzyc pliki pomocnicze!')
    
    def create_excel_file(self) -> None:
        """
            funkcja tworzy nowy plik excela -- BEZ MIESIACA - TODO:
        """
        print('Zaczynam tworzyc plik ewidencji')
        try:
            self._get_data_to_csv_files()
            print(f'\tTworze plik {self.output_path}...')
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                stan_form = '='
                csv_dir_data = os.listdir(self.csv_dir)
                for file in tqdm(csv_dir_data, desc='\tTworze plik glowny: ', unit=' jednostek'):
                    filename = file.removesuffix('.csv')
                    stan_form += f"'{filename}'.D2+"
                    data = pd.read_csv(f'{self.csv_dir}{file}', sep=';')
                    data.to_excel(writer, sheet_name=filename, index=False)
                    sheet = writer.sheets[filename]
                    sheet.cell(1, 4, 'STAN') # opis formuly
                    sheet.cell(1, 4).fill = self.CELL_COLOR
                    sheet.cell(1, 4).border = self.BORDER_STYLE
                    end_cell = len(data['opis']) + 2 
                    sheet.cell(2, 4, f'=SUM(B2:B{end_cell})-SUM(C2:C{end_cell})')
                    sheet.cell(2, 4).fill = self.CELL_COLOR
                print('Tworzenie arkuszy ukonczone pomyslnie.\nTworze arkusz STAN...')
                sheet = writer.book.create_sheet('STAN')
                sheet.cell(1, 1, 'STAN')
                sheet.cell(1, 1).fill = self.CELL_COLOR
                #stan_form = stan_form.removesuffix('+')
                sheet.cell(1, 2, stan_form)
                sheet.cell(1, 3, 'jezeli nie dziala - usunac i dodac jakikolwiek znak, zeby sie przeformatowalo')#FIXME
                print('Pomyslnie ukonczono tworzenie pliku.')
        except Exception as e:
            raise ValueError(f'Nie mozna stworzyc pliku {self.output_path}.\nBlad: {e}')

    def _get_previos_output_length(self, sheet_name: str) -> int:
        """
            sheet_name: str -> nazwa arkusza, ktorego aktualna dlugosc trzeba sprawdzic
            return:
                (int) dlugosc opisu danego arkusza lub 0 jezeli nie ma
        """
        print(f'Pobieram poprzednia dlugosc arkusza {sheet_name}')
        try:
            with pd.ExcelFile(self.output_path, 'openpyxl',) as file:
                if sheet_name in file.sheet_names: 
                    excel_df = pd.read_excel(file, sheet_name=sheet_name)
                    #print(excel_df)
                    return len(excel_df['opis'])
        except Exception as e:
            raise ValueError(f'Nie udalo sie odczytac poprzedniej dlogosci arkusza.\nBlad: {e}')
        return 0

    def rewrite_STAN_formula(self) -> None:
        """
            funkcja pisze od nowa formule w arkuszu STAN (zliczajaca sumy wszystkich jednostek)
        """
        print('Tworze nowa formule w arkuszu STAN...')
        try:
            with pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                list_with_sheet_names = [sheet for sheet in writer.sheets]
                list_with_sheet_names.remove('STAN')
                stan_form = '=0'
                for sheet in list_with_sheet_names:
                    stan_form += f"+'{sheet}'.D2"
                writer.sheets['STAN'].cell(1, 2, stan_form)
            print('Poprawnie utworzono nowa formule.')
        except Exception as e:
            print(f'Nie udalo sie stworzyc nowej formuly.\nBlad:{e}')


    def write_data_to_one_excel_file(self) -> None:
        """
            funkcja dopisuje kolejny miesiac do aktualnego pliku z ewidencji
        """
        self._get_data_to_csv_files()
        month = input('Podaj miesiac, ktorego dotyczy wyciag: ')
        csv_dir_data = os.listdir(self.csv_dir)
        if len(csv_dir_data) == 0:
            raise FileNotFoundError('Nie ma plikow do przetowrzenia')
        for csv_file in tqdm(csv_dir_data, desc='Przetwarzam arkusze: ', unit=' arkusze'):
            csv_filename = csv_file.removesuffix('.csv')
            csv_df = pd.read_csv(f'{self.csv_dir}/{csv_file}', sep=';')

            previous_data_length = self._get_previos_output_length(csv_filename)

            with pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if csv_filename in writer.book.sheetnames:
                    sheet = writer.book[csv_filename]
                else:
                    sheet = writer.book.create_sheet(csv_filename)

                sheet.cell(previous_data_length + 2, 4, month)
                sheet.cell(previous_data_length + 2, 4).fill = self.CELL_COLOR
                
                current_data_length = previous_data_length + len(csv_df['opis']) + 2
                sheet.cell(2, 4, f'=SUM(B2:B{current_data_length})')
                csv_df.to_excel(writer, sheet_name=csv_filename, index=False, header=False, startrow=previous_data_length+2)
        
if __name__ == '__main__':
    #filepath = input('filename: ')
    test = ExcelHandler('../dane-ewidencja/03_27_2024_10_52_07-RB 2024 01 (3).xlsx')
    #test.get_data_to_csv_files()
    test.create_excel_file()
    test.write_data_to_one_excel_file()
    test.rewrite_STAN_formula()